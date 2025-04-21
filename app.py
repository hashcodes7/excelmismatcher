import streamlit as st
import pandas as pd
import warnings
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

warnings.filterwarnings('ignore')
def add_row_to_dataframe(df, new_row):
    if df is None:
        # Create a new DataFrame if it doesn't exist
        df = new_row
    else:
        # Append the new row to the existing DataFrame
        df = pd.concat([df, new_row], ignore_index=True)
    return df
def index_to_column_name(index):
    col_str = ''
    while index >= 0:
        col_str = chr(index % 26 + 65) + col_str
        index = index // 26 - 1
    return col_str

status=''
st.header("Sheet view") 
st.sidebar.write('v 1.0.1 . For Cognizant Levi Team Internal Use only.')
xlfile1=None
xlfile2=None
xlfile1 = st.sidebar.file_uploader("Upload first Excel file", type=['xlsx'], key='file1')
xlfile2 = st.sidebar.file_uploader("Upload second Excel file", type=['xlsx'], key='file2')
if xlfile1 and xlfile2:
    xls1 = pd.ExcelFile(xlfile1)
    xls2 = pd.ExcelFile(xlfile2)
    if xlfile1 and xlfile2:
        xls = pd.ExcelFile(xlfile1)
        common_sheets = list(set(xls1.sheet_names) & set(xls2.sheet_names))
        if len(common_sheets)==0:
            st.write("No common sheets found in both Excel files.")
            st.write("Sheets in Excel 1:", xls1.sheet_names)
            st.write("Sheets in Excel 2:", xls2.sheet_names)
        else:
            sheet_selector = st.sidebar.selectbox("Select sheet:", common_sheets)   
            header = st.sidebar.number_input("Header (row number)", min_value=0, value=0, step=1)
            df1 = pd.read_excel(xlfile1, sheet_name=sheet_selector,header=header)
            df2 = pd.read_excel(xlfile2, sheet_name=sheet_selector,header=header)
            UniqueRow = st.sidebar.selectbox("Primary Key/ Unique Value column", df1.columns) 

            st.markdown(f"### Currently Selected: `{sheet_selector}`")
            df1.index = df1.index + 1
            df2.index = df2.index + 1
            st.write(df1)
            st.write(df2)
            if UniqueRow:
                # columns = [UniqueRow] + [col for col in df1.columns if col != UniqueRow]
                # df1 = df1[columns]
                # df2=df2[columns]
                # df1=df1[[UniqueRow]+[c for c in df1 if c !=UniqueRow]]
                # df2=df2[[UniqueRow]+[c for c in df2 if c !=UniqueRow]]


                #line below is error handling in case one file is empty, doesnt has unique row, or wrong header is selected
                if UniqueRow not in df1.columns or UniqueRow not in df2.columns:
                    print("The unique row",UniqueRow,"could not be found in one of the sheets, please check the headers and if they contain unique row")
                    print("current file headers for file 1:",df1.columns,"\ncurrent file headers for file 2:",df2.columns)
                    quit()
                uid1list = df1[UniqueRow].tolist()
                uid2list = df2[UniqueRow].tolist()
                set1 = set(uid1list)
                set2 = set(uid2list)

                unique_elements = set1.symmetric_difference(set2)
                unique_from_list1 = unique_elements.intersection(set1)
                unique_from_list2 = unique_elements.intersection(set2)
                onlyonefiledataframe=None
                for uniques1 in unique_from_list1:
                    ind1=uid1list.index(uniques1)
                    row_df1 = df1.iloc[[ind1]].copy()
                    row_df1['source']=xlfile1.name
                    onlyonefiledataframe = add_row_to_dataframe(onlyonefiledataframe, row_df1)
                    onlyonefiledataframe['source']=xlfile1.name
                # empty line
                if onlyonefiledataframe is not None and not onlyonefiledataframe.empty:
                    onlyonefiledataframe.loc[len(onlyonefiledataframe)] = [None] * len(onlyonefiledataframe.columns)

                #...........
                for uniques2 in unique_from_list2:
                    ind2=uid2list.index(uniques2)
                    row_df2 = df2.iloc[[ind2]].copy()
                    row_df2['source']=xlfile2.name
                    onlyonefiledataframe = add_row_to_dataframe(onlyonefiledataframe, row_df2)
                
                #----------------------------------
                if onlyonefiledataframe is not None:
                    text = f'''Unique Row({UniqueRow})'''
                    onlyonefiledataframe.insert(0, text, onlyonefiledataframe[UniqueRow])
                if onlyonefiledataframe is not None:
                    col = onlyonefiledataframe.pop('source')
                    onlyonefiledataframe.insert(0, 'source', col)

                #the next program creates second sheet for checking values which are different, but having same unique identifier
                df = None
                globalmismatched=[]
                for uid in uid1list:
                    if uid in uid2list:
                        ind1=uid1list.index(uid)
                        ind2=uid2list.index(uid)
                        row_df1 = df1.iloc[[ind1]].copy()
                        row_df2 = df2.iloc[[ind2]].copy()
                        # are_not_equal = not row_df1.equals(row_df2)
                        # if are_not_equal==True:
                        differing_columns = []
                        column_ids=[]
                        # print(row_df1)
                        for col in row_df1.columns:
                            val1 = row_df1[col].values[0]
                            val2 = row_df2[col].values[0]
                            if pd.isna(val1) and pd.isna(val2):
                                continue  # Both are NaN, so they are considered equal
                            if val1 != val2:
                                differing_columns.append(col)
                        if len(differing_columns)>0:
                            row_df1.insert(1, 'Unmatched Column', '')
                            for mismatch in differing_columns:
                                column_ids.append(index_to_column_name(row_df1.columns.get_loc(mismatch)+2))
                            # print(differing_columns,len(differing_columns))

                            text = f'''Unique Row({UniqueRow})'''
                            row_df1.insert(0, text, row_df1[UniqueRow])
                            row_df1.insert(1, 'Comments', f"{len(differing_columns)} mismatched")
                            combined = [f"{a}({b})" for a, b in zip(column_ids, differing_columns)]
                            comments = ', '.join(combined)
                            print(comments)
                            row_df1["Unmatched Column"]=comments
                            for mismatch in column_ids:
                                if df is None:
                                    globalmismatched.append(f'{mismatch}{2}')
                                else:
                                    height = df.shape[0]
                                    globalmismatched.append(f'{mismatch}{height+2}')

                            # row_df2['Unmatched Column(column id)'] = comments
                            row_df1['-----------------'] = "-----------------"
                            row_df1['------------------'] = "------------------"
                            row_df1 = row_df1.reset_index(drop=True)
                            row_df2 = row_df2.reset_index(drop=True)
                            rowdf = pd.concat([row_df1, row_df2], axis=1)
                            df = add_row_to_dataframe(df, rowdf)
                          
                # columns = [UniqueRow] + [col for col in df.columns if col != UniqueRow]
                # print(UniqueRow,type(UniqueRow))
                # df= df[columns]

                    #error handling otherwise if same files are compared these dataframes are never made and object type stays None empty causing 
                    # error 'NoneType' object has no attribute 'to_excel'
                # if onlyonefiledataframe is not None:
                # with pd.ExcelWriter('combined.xlsx') as writer:
                #     onlyonefiledataframe.to_excel(writer, sheet_name='Missing', index=False)
                #     df.to_excel(writer, sheet_name='MisMatched', index=False)
                if onlyonefiledataframe is not None and df is not None:
                    status="There are both missing and mismatched columns"
                    with pd.ExcelWriter('combined.xlsx') as writer:
                        onlyonefiledataframe.to_excel(writer, sheet_name='Missing', index=False)
                        df.to_excel(writer, sheet_name='MisMatched', index=False)
                elif onlyonefiledataframe is not None and df is None:
                    status="There are no Mismatched columns"
                    with pd.ExcelWriter('combined.xlsx') as writer:
                        onlyonefiledataframe.to_excel(writer, sheet_name='Missing', index=False)
                        # df.to_excel(writer, sheet_name='MisMatched', index=False)
                elif onlyonefiledataframe is None and df is not None:
                    status="There are no Missing columns"
                    with pd.ExcelWriter('combined.xlsx') as writer:
                        # onlyonefiledataframe.to_excel(writer, sheet_name='Missing', index=False)
                        df.to_excel(writer, sheet_name='MisMatched', index=False)
                else:
                    status="There are no missing and mismatched columns, both sheets are exactly same."
                    
                ##########################################
                if (status!=''):
                    st.sidebar.write(status)
                if len(globalmismatched)!=0:
                    wb = load_workbook('combined.xlsx')
                    ws = wb['MisMatched']
                    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

                    # Apply the yellow fill to all cells in alldiff
                    print(globalmismatched)
                    for cell in globalmismatched:
                        ws[cell].fill = yellow_fill
                    wb.save('combined.xlsx')
                
                
                # Download button for the combined Excel file
                if df is not None and onlyonefiledataframe is not None:
                    with open('combined.xlsx', 'rb') as f:
                        st.sidebar.download_button(
                            label="Download Excel file",
                            data=f,
                            file_name='combined.xlsx',
                            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                        )
st.sidebar.write('For any changes or mantainence reach out to Levi Team')

                
                

